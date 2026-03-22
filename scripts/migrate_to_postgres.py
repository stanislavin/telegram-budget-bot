#!/usr/bin/env python3
"""Migrate historical expense data from Excel files into Neon PostgreSQL.

Usage:
    python scripts/migrate_to_postgres.py              # normal run (aborts if table has data)
    python scripts/migrate_to_postgres.py --force      # truncate first, then load
"""

import argparse
import asyncio
import os
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

# Add project root to path so we can import util.*
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from util.config import DATABASE_URL
from util.postgres import _clean_dsn

import asyncpg

BUDGETS_DIR = Path('tmp_budget_files/budgets')
ACTIVE_FILE = Path('tmp_budget_files/budgeting101.xlsx')
SHEET_NAME = 'Form Responses 1'


def parse_timestamp(value) -> datetime | None:
    """Parse a timestamp from either a datetime object or string."""
    if isinstance(value, datetime):
        return value
    if not isinstance(value, str) or not value.strip():
        return None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%m/%d/%Y %H:%M:%S', '%d/%m/%Y %H:%M:%S'):
        try:
            return datetime.strptime(value.strip(), fmt)
        except ValueError:
            continue
    return None


def read_expenses_from_xlsx(filepath: Path) -> list[tuple]:
    """Read expenses from an xlsx file's 'Form Responses 1' sheet.

    Returns list of (timestamp, amount, currency, category, description, source_file).
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"  Skipping {filepath.name}: no '{SHEET_NAME}' sheet")
        wb.close()
        return []

    ws = wb[SHEET_NAME]
    rows = []
    source = filepath.name

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header

        # Minimum: timestamp (A), amount (B), category (C)
        if not row or len(row) < 3:
            continue

        ts = parse_timestamp(row[0])
        if ts is None:
            continue

        try:
            amount = float(str(row[1]))
        except (TypeError, ValueError):
            continue

        category = str(row[2]).strip() if row[2] else ''
        if not category:
            continue

        description = str(row[3]).strip() if len(row) > 3 and row[3] else ''

        # Currency: column F (index 5) if present, else default to 'RSD'
        currency = 'RSD'
        if len(row) > 5 and row[5]:
            currency = str(row[5]).strip()

        rows.append((ts, amount, currency, category, description, source))

    wb.close()
    return rows


def deduplicate(records: list[tuple]) -> list[tuple]:
    """Remove duplicates based on (timestamp, amount, category, description, currency).

    First occurrence wins (monthly files loaded before active file).
    """
    seen = set()
    unique = []
    for rec in records:
        key = (rec[0], rec[1], rec[2], rec[3], rec[4])  # ts, amount, currency, category, description
        if key not in seen:
            seen.add(key)
            unique.append(rec)
    return unique


async def main():
    parser = argparse.ArgumentParser(description='Migrate Excel data to PostgreSQL')
    parser.add_argument('--force', action='store_true', help='Truncate table before inserting')
    args = parser.parse_args()

    if not DATABASE_URL:
        print('ERROR: DATABASE_URL is not set.')
        sys.exit(1)

    dsn = _clean_dsn(DATABASE_URL)
    conn = await asyncpg.connect(dsn)

    # Safety check
    count = await conn.fetchval('SELECT COUNT(*) FROM expenses')
    if count > 0 and not args.force:
        print(f'ERROR: Table already has {count} rows. Use --force to truncate first.')
        await conn.close()
        sys.exit(1)

    if args.force and count > 0:
        print(f'Truncating {count} existing rows...')
        await conn.execute('TRUNCATE TABLE expenses RESTART IDENTITY')

    # 1. Read monthly files (sorted alphabetically = chronological)
    all_records = []
    if BUDGETS_DIR.exists():
        monthly_files = sorted(BUDGETS_DIR.glob('*.xlsx'))
        print(f'Found {len(monthly_files)} monthly files in {BUDGETS_DIR}')
        for f in monthly_files:
            rows = read_expenses_from_xlsx(f)
            print(f'  {f.name}: {len(rows)} rows')
            all_records.extend(rows)
    else:
        print(f'Warning: {BUDGETS_DIR} not found, skipping monthly files')

    # 2. Read active file
    if ACTIVE_FILE.exists():
        rows = read_expenses_from_xlsx(ACTIVE_FILE)
        print(f'  {ACTIVE_FILE.name}: {len(rows)} rows')
        all_records.extend(rows)
    else:
        print(f'Warning: {ACTIVE_FILE} not found, skipping')

    # 3. Deduplicate
    before = len(all_records)
    all_records = deduplicate(all_records)
    after = len(all_records)
    print(f'\nDeduplication: {before} -> {after} records ({before - after} duplicates removed)')

    # 4. Bulk insert
    if all_records:
        await conn.copy_records_to_table(
            'expenses',
            records=all_records,
            columns=['timestamp', 'amount', 'currency', 'category', 'description', 'source_file'],
        )
        print(f'\nInserted {len(all_records)} records into expenses table.')
    else:
        print('\nNo records to insert.')

    # Verify
    final_count = await conn.fetchval('SELECT COUNT(*) FROM expenses')
    print(f'Final row count: {final_count}')

    await conn.close()


if __name__ == '__main__':
    asyncio.run(main())
